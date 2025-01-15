import tkinter as tk
from tkinter import ttk
import openpyxl
from openpyxl import Workbook


File_Name = "classwork.xlsx"

root = tk.Tk()
root.title("Data Store System")
root.geometry("700x500")


def initializeExcel():
    try:
        print("File Found !")
        workbook = openpyxl.load_workbook(File_Name)
        worksheet=workbook.active
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            table.insert("",tk.END,values=row)
            
    except Exception as e:
        print(e)
        print('File Not Found !')
        workbook = Workbook()
        worksheet=workbook.active
        worksheet.append(["Name", "age"])
        print("File Created Automatically !")
        workbook.save(File_Name)



ui_frame = tk.Frame(root)
ui_frame.grid(row=1, column=0)

tk.Label(ui_frame, text="Name").grid(row=0, column=0)
entryName=tk.Entry(ui_frame)
entryName.grid(row=0, column=1)

tk.Label(ui_frame, text="Age").grid(row=1, column=0)
entryAge=tk.Entry(ui_frame)
entryAge.grid(row=1, column=1)

table = ttk.Treeview(ui_frame, column=("Name", "Age"), show="headings")
table.heading("Name", text="Name")
table.heading("Age", text="Age")
table.grid(row=4, column=0, columnspan=2)

initializeExcel()
# Lambda Table Insertion Function.
tableInsert = lambda **kwargs:table.insert("", tk.END, values=(kwargs['name'],kwargs['age']))


def save_data():
    print("Logic Here !")
    name=entryName.get()
    age=entryAge.get()
    print(name, age)

    # Table enter
    tableInsert(name=name, age=age)
    # Insert Current Data Into The File.
    workbook = openpyxl.load_workbook(File_Name)
    worksheet = workbook.active
    worksheet.append((name, age))
    workbook.save(File_Name)

# Create Save Button
tk.Button(ui_frame, text="save", command=save_data).grid(row=2, column=1)

# Keep The Program In Running State.
root.mainloop()